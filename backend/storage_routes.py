import io
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from auth_utils import require_admin
from database import storage_settings_collection, storage_cache_collection
from storage_status import refresh_storage_status, _tier_for
from activity_service import log_activity

router = APIRouter(prefix="/api/storage", tags=["storage"])


@router.get("/status")
async def get_storage_status(admin: dict = Depends(require_admin)):
    """Cached snapshot (refreshed daily by the scheduler, or on-demand via
    /refresh) — doesn't recompute live on every dashboard load."""
    snapshot = await storage_cache_collection.find_one({"_id": "latest"}, {"_id": 0})
    if not snapshot:
        snapshot = refresh_storage_status()
        snapshot.pop("_id", None)
    return snapshot


@router.post("/refresh")
async def refresh_storage(admin: dict = Depends(require_admin)):
    """Force a live recompute now (admin's "Refresh Now" button)."""
    snapshot = refresh_storage_status()
    snapshot.pop("_id", None)
    return snapshot


class StorageSettingsIn(BaseModel):
    account_email: Optional[str] = None
    plan: Optional[str] = None
    limit_mb: Optional[float] = None
    used_mb: Optional[float] = None


@router.put("/settings/{service}")
async def update_storage_settings(service: str, data: StorageSettingsIn, admin: dict = Depends(require_admin)):
    """Admin-entered figures for services with no billing API wired up
    (render, netlify), or a limit override once mongodb is upgraded past M0."""
    if service not in ("render", "netlify", "mongodb"):
        raise HTTPException(status_code=400, detail="Unknown service")
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    await storage_settings_collection.update_one({"_id": service}, {"$set": updates}, upsert=True)
    snapshot = refresh_storage_status()
    snapshot.pop("_id", None)
    return snapshot


@router.post("/mark-upgraded/{service}")
async def mark_upgraded(service: str, new_limit_mb: float, admin: dict = Depends(require_admin)):
    """Convenience for the "Mark as Upgraded" fallback button — bumps the
    recorded limit and immediately re-checks thresholds against it."""
    if service not in ("render", "netlify", "mongodb"):
        raise HTTPException(status_code=400, detail="Unknown service")
    await storage_settings_collection.update_one(
        {"_id": service}, {"$set": {"limit_mb": new_limit_mb}}, upsert=True,
    )
    snapshot = refresh_storage_status()
    snapshot.pop("_id", None)
    return snapshot


TFD_NAVY = "024396"
TIER_LABELS = {"ok": "OK", "warning": "WARNING", "critical": "CRITICAL"}


@router.get("/report")
async def download_storage_report(admin: dict = Depends(require_admin)):
    """On-demand branded Excel export of the current storage snapshot —
    summary, per-service category breakdown, and active warnings."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    snapshot = await storage_cache_collection.find_one({"_id": "latest"}, {"_id": 0})
    if not snapshot:
        snapshot = refresh_storage_status()
        snapshot.pop("_id", None)

    header_fill = PatternFill(start_color=TFD_NAVY, end_color=TFD_NAVY, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14, color=TFD_NAVY)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "The Financial Doctor — Storage & Infrastructure Report"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')}"
    ws["A3"] = f"Snapshot last refreshed: {snapshot.get('updated_at', '')}"

    ws.append([])
    header_row = ["Service", "Used", "Limit", "% Used", "Status", "Source", "Account", "Plan"]
    ws.append(header_row)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font

    warnings = []
    for svc in snapshot.get("services", {}).values():
        tier = _tier_for(svc.get("percent", 0))
        status_label = TIER_LABELS.get(tier, "OK")
        if tier != "ok":
            warnings.append(f"{svc['label']}: {status_label} ({svc['percent']}% used)")
        ws.append([
            svc["label"],
            f"{svc['used_mb']:.1f} MB" if svc.get("used_mb", 0) < 1024 else f"{svc['used_mb']/1024:.2f} GB",
            f"{svc['limit_mb']:.0f} MB" if svc.get("limit_mb", 0) < 1024 else f"{svc['limit_mb']/1024:.1f} GB",
            f"{svc.get('percent', 0)}%",
            status_label,
            "Live" if svc.get("live") else "Estimated (manual entry)",
            svc.get("account_email") or "",
            svc.get("plan") or "",
        ])

    ws.append([])
    ws.append(["Active Warnings"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, color="B3261E")
    if warnings:
        for w in warnings:
            ws.append([w])
    else:
        ws.append(["None — all services within normal range."])

    for col, width in zip("ABCDEFGH", [22, 12, 12, 10, 12, 22, 24, 16]):
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("Category Breakdown")
    ws2.append(["Service", "Category", "Size"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for svc in snapshot.get("services", {}).values():
        for b in svc.get("breakdown", []):
            size = f"{b['size_mb']:.2f} MB" if b["size_mb"] < 1024 else f"{b['size_mb']/1024:.2f} GB"
            ws2.append([svc["label"], b["category"], size])
    for col, width in zip("ABC", [20, 45, 14]):
        ws2.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"tfd_storage_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    await log_activity(admin["sub"], "report_downloaded", "Downloaded the storage & infrastructure report", link="/portal/admin/storage")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
