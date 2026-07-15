"""Lead Management CRM routes.

Admin:
  POST   /api/leads/          — create a lead
  GET    /api/leads/          — list all leads (filter by status, assigned_to)
  GET    /api/leads/{id}      — single lead detail
  PUT    /api/leads/{id}      — update lead fields
  DELETE /api/leads/{id}      — delete a lead
  POST   /api/leads/{id}/assign  — assign lead to an employee
  GET    /api/leads/stats     — lead pipeline stats

Employee:
  GET    /api/leads/my        — leads assigned to current employee
  POST   /api/leads/{id}/status — update lead status + follow-up note
"""

import io
import logging
import random
import re
import uuid
from datetime import datetime, timezone, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from pydantic import BaseModel, Field

from lead_models import LeadCreate, LeadUpdate, LeadStatusUpdate, LeadNameUpdate, LeadAlternatePhoneUpdate, LeadInDB, LeadOut, CallOutcomeIn, TransferIn
from notification_models import NotificationInDB
from auth_utils import get_current_user_payload, require_admin
from database import leads_collection, users_collection, db, reminders_collection, pipelines_collection, lead_batches_collection, notifications_collection
from notification_service import create_notification
from activity_service import log_activity
from email_service import send_career_application_received_email

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/leads", tags=["leads"])

web_leads_collection = db["web_leads"]
career_leads_collection = db["career_leads"]

VALID_STATUSES = {"new", "contacted", "follow_up", "interested", "converted", "lost"}
STATUS_LABELS = {
    "new": "New", "contacted": "Contacted", "follow_up": "Follow Up",
    "interested": "Interested", "converted": "Converted", "lost": "Lost",
}


def to_lead_out(doc: dict) -> LeadOut:
    # Omit fields the doc doesn't have (or has as None) instead of passing
    # None explicitly — lets Pydantic fall back to each field's own default
    # (e.g. reassign_count: int = 0) rather than failing validation on
    # older lead documents that predate a newer field being added.
    data = {k: doc[k] for k in LeadOut.model_fields if doc.get(k) is not None}
    return LeadOut(**data)


async def notify_admins_employee_added_lead(emp_id: str, emp_name: str, lead_name: str, lead_phone: str) -> None:
    """In-app-only (no push) notification to every admin that an employee
    self-added a lead. Multiple leads added by the same employee collapse
    into one growing notification instead of flooding the bell with a
    separate row per lead — it keeps updating the same unread entry until
    an admin reads it, then the next add starts a fresh one."""
    admins = users_collection.find({"role": "admin"})
    async for adm in admins:
        existing = await notifications_collection.find_one({
            "user_id": adm["id"], "type": "employee_lead_added",
            "read": False, "meta.employee_id": emp_id,
        })
        if existing:
            count = (existing.get("meta") or {}).get("count", 1) + 1
            await notifications_collection.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "body": f"{emp_name} added {count} leads themselves — latest: {lead_name} ({lead_phone})",
                    "created_at": datetime.now(timezone.utc),
                    "meta": {"employee_id": emp_id, "count": count},
                }},
            )
        else:
            note = NotificationInDB(
                user_id=adm["id"], title="Employee Added a Lead",
                body=f"{emp_name} added {lead_name} ({lead_phone}) themselves",
                type="employee_lead_added", link="/portal/admin/leads",
                meta={"employee_id": emp_id, "count": 1},
            )
            await notifications_collection.insert_one(note.dict())


async def find_pipeline_for_employee(employee_id: str) -> Optional[str]:
    """Returns the id of the (first) active pipeline assigned to this
    employee, or None if they have no pipeline. Used to auto-attach a
    pipeline to a lead the moment it's assigned to someone."""
    if not employee_id:
        return None
    pipeline = await pipelines_collection.find_one(
        {"assigned_to": employee_id, "is_active": {"$ne": False}}
    )
    return pipeline["id"] if pipeline else None


# ── Admin: CRUD ──────────────────────────────────────────────────

@router.post("/", response_model=LeadOut)
async def create_lead(data: LeadCreate, admin: dict = Depends(require_admin)):
    if not data.assigned_to:
        raise HTTPException(status_code=400, detail="Assign this lead to yourself or an employee")
    lead = LeadInDB(**data.model_dump())
    emp = await users_collection.find_one({"id": data.assigned_to})
    if not emp:
        raise HTTPException(status_code=404, detail="Selected employee not found")
    lead.assigned_to_name = emp.get("name")
    if not lead.pipeline_id:
        lead.pipeline_id = await find_pipeline_for_employee(data.assigned_to)
    await leads_collection.insert_one(lead.model_dump())
    await create_notification(
        user_id=data.assigned_to,
        title="New Lead Assigned",
        body=f"{lead.name} ({lead.phone}) — {lead.service_interest or 'General'}",
        n_type="lead",
        link=f"/portal/employee/leads?leadId={lead.id}",
    )
    await log_activity(
        admin["sub"], "lead_added",
        f"Added a lead: {lead.name} ({lead.phone}) — assigned to {lead.assigned_to_name}",
        link="/portal/admin/leads",
    )
    return to_lead_out(lead.model_dump())


@router.get("/stats")
async def lead_stats(admin: dict = Depends(require_admin)):
    pipeline = [
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
    ]
    status_counts = {}
    async for doc in leads_collection.aggregate(pipeline):
        status_counts[doc["_id"]] = doc["count"]

    total = sum(status_counts.values())

    source_pipeline = [
        {"$group": {"_id": "$source", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    source_counts = []
    async for doc in leads_collection.aggregate(source_pipeline):
        source_counts.append({"source": doc["_id"] or "unknown", "count": doc["count"]})

    return {
        "total": total,
        "by_status": status_counts,
        "by_source": source_counts,
    }


DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


@router.get("/my", response_model=list[LeadOut])
async def my_leads(date: Optional[str] = Query(default=None), payload: dict = Depends(get_current_user_payload)):
    """Leads assigned to the current employee. Leads that have already been
    called at least once (call_touched), that are a referral from an
    existing client, or that the employee added themselves, always show —
    a referral is a warm lead and a self-added lead was just entered by the
    employee, so neither should get buried behind a pile of cold ones.
    Everything else (brand-new, never-called, non-referral leads — however
    admin added them: manual, Excel batch, shuffle-assign) is capped to 10
    at a time — the 11th only appears once one of the current 10 has been
    called. Pass `date` (YYYY-MM-DD, see /my/dates) to work through a
    specific day's leads instead of always the oldest first — e.g. calling
    a batch admin uploaded 10 days ago, or jumping straight to today's."""
    if date is not None and not DATE_RE.match(date):
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")

    touched_cursor = leads_collection.find(
        {"assigned_to": payload["sub"], "$or": [
            {"call_touched": True},
            {"referred_by_lead_id": {"$exists": True, "$ne": None}},
            {"source": "employee_added"},
        ]}
    ).sort("updated_at", -1)
    touched = [doc async for doc in touched_cursor]

    # Querying a field as None matches both "explicitly null" and "missing
    # entirely" in MongoDB, so this alone covers old leads (field never
    # existed) and new non-referral leads (field explicitly None) in one go.
    fresh_query = {
        "assigned_to": payload["sub"], "call_touched": {"$ne": True},
        "referred_by_lead_id": None, "source": {"$ne": "employee_added"},
    }
    if date:
        fresh_query["created_at"] = {"$regex": f"^{date}"}
    fresh_cursor = leads_collection.find(fresh_query).sort("created_at", 1).limit(10)
    fresh = [doc async for doc in fresh_cursor]

    return [to_lead_out(doc) for doc in touched + fresh]


@router.get("/my/dates")
async def my_lead_dates(payload: dict = Depends(get_current_user_payload)):
    """Distinct calendar dates (from created_at) among the employee's
    not-yet-called leads, newest first with counts — powers a picker so they
    can jump straight to a specific admin-added batch/day instead of only
    working the global oldest-first queue."""
    pipeline = [
        {"$match": {
            "assigned_to": payload["sub"],
            "call_touched": {"$ne": True},
            "referred_by_lead_id": None,
            "source": {"$ne": "employee_added"},
        }},
        {"$group": {"_id": {"$substrCP": ["$created_at", 0, 10]}, "count": {"$sum": 1}}},
        {"$sort": {"_id": -1}},
    ]
    dates = []
    async for doc in leads_collection.aggregate(pipeline):
        dates.append({"date": doc["_id"], "count": doc["count"]})
    return dates


@router.get("/my/summary")
async def my_leads_summary(payload: dict = Depends(get_current_user_payload)):
    """Counts for the employee's own queue — lets the UI tell them how many
    more never-called leads are waiting behind the 10-at-a-time cap in /my,
    so a freshly assigned batch doesn't look like it silently vanished."""
    total_new = await leads_collection.count_documents(
        {"assigned_to": payload["sub"], "call_touched": {"$ne": True}}
    )
    visible_new = min(total_new, 10)
    return {"total_new": total_new, "visible_new": visible_new, "queued": max(total_new - visible_new, 0)}


@router.get("/", response_model=list[LeadOut])
async def list_leads(
    status: Optional[str] = None,
    assigned_to: Optional[str] = None,
    search: Optional[str] = None,
    pipeline_id: Optional[str] = None,
    limit: int = Query(default=200, le=500),
    admin: dict = Depends(require_admin),
):
    query = {}
    if status and status in VALID_STATUSES:
        query["status"] = status
    if assigned_to:
        query["assigned_to"] = assigned_to
    if pipeline_id:
        query["pipeline_id"] = pipeline_id
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
        ]
    cursor = leads_collection.find(query).sort("updated_at", -1).limit(limit)
    return [to_lead_out(doc) async for doc in cursor]


@router.get("/website")
async def get_website_leads(admin: dict = Depends(require_admin)):
    cursor = web_leads_collection.find().sort("created_at", -1)
    leads = []
    async for doc in cursor:
        doc.pop("_id", None)
        leads.append(doc)
    return leads


@router.get("/career")
async def get_career_leads(admin: dict = Depends(require_admin)):
    cursor = career_leads_collection.find().sort("created_at", -1)
    leads = []
    async for doc in cursor:
        doc.pop("_id", None)
        leads.append(doc)
    return leads


@router.get("/search")
async def search_all_leads(q: str = Query(..., min_length=2), payload: dict = Depends(get_current_user_payload)):
    """Global search by (partial) name or phone number — any employee can check if a
    lead already exists before calling, so no one duplicates outreach."""
    cursor = leads_collection.find({
        "$or": [
            {"name": {"$regex": q, "$options": "i"}},
            {"phone": {"$regex": q, "$options": "i"}},
        ]
    }).limit(20)
    results = [to_lead_out(doc) async for doc in cursor]
    return results


@router.get("/batches")
async def list_lead_batches(admin: dict = Depends(require_admin)):
    """Returns list of lead upload batches with stats per batch."""
    pipeline_agg = [
        {"$match": {"batch_id": {"$exists": True, "$ne": None}}},
        {"$group": {
            "_id": "$batch_id",
            "date": {"$first": "$batch_date"},
            "total": {"$sum": 1},
            "interested": {"$sum": {"$cond": [{"$eq": ["$status", "interested"]}, 1, 0]}},
            "converted": {"$sum": {"$cond": [{"$eq": ["$status", "converted"]}, 1, 0]}},
            "follow_up": {"$sum": {"$cond": [{"$eq": ["$status", "follow_up"]}, 1, 0]}},
            "lost": {"$sum": {"$cond": [{"$eq": ["$status", "lost"]}, 1, 0]}},
            "called": {"$sum": {"$cond": [{"$eq": ["$call_touched", True]}, 1, 0]}},
        }},
        {"$sort": {"date": -1}},
        {"$limit": 50},
    ]
    batches = []
    async for doc in leads_collection.aggregate(pipeline_agg):
        meta = await lead_batches_collection.find_one({"batch_id": doc["_id"]}) or {}
        batches.append({
            "batch_id": doc["_id"],
            "date": doc["date"],
            "total": doc["total"],
            "called": doc["called"],
            "interested": doc["interested"],
            "converted": doc["converted"],
            "follow_up": doc["follow_up"],
            "lost": doc["lost"],
            # Import-time stats, kept around so admin can check them later too,
            # not just in the one-time success alert.
            "incomplete_count": meta.get("incomplete_count", 0),
            "duplicate_infile_count": meta.get("duplicate_infile_count", 0),
            "duplicate_existing_count": meta.get("duplicate_existing_count", 0),
            "reassigned_existing_count": meta.get("reassigned_existing_count", 0),
            "included_existing_duplicates": meta.get("included_existing_duplicates", False),
        })
    return batches


@router.get("/batches/{batch_id}")
async def get_batch_leads(
    batch_id: str,
    employee_filter: Optional[str] = None,
    admin: dict = Depends(require_admin),
):
    """Returns all leads in a batch, optionally filtered by assigned employee."""
    query = {"batch_id": batch_id}
    if employee_filter:
        query["assigned_to"] = employee_filter
    cursor = leads_collection.find(query).sort("assigned_to_name", 1)
    return [to_lead_out(doc) async for doc in cursor]


REASON_LABELS = {
    "npc": "no response (NPC)", "switchoff": "switched off",
    "network_issue": "network issue", "busy": "busy", "invalid": "invalid number",
}
SUB_STAGE_LABELS = {
    **REASON_LABELS,
    "interested": "Interested", "follow_up": "Follow Up", "not_interested": "Not Interested",
    "converted": "Converted", "lost": "Lost (on call)",
}


def lead_call_outcome_label(lead: dict) -> str:
    """Human label for a lead's most recent call result — Interested / Not
    Interested / Switched Off / Invalid Number / Busy / No Response /
    Network Issue / Converted / Lost / Not called yet — for the per-lead
    row in the Excel report."""
    if not lead.get("call_touched"):
        return "Not called yet"
    for h in reversed(lead.get("status_history", [])):
        if h.get("sub_stage"):
            return SUB_STAGE_LABELS.get(h["sub_stage"], h["sub_stage"])
    return "Connected" if lead.get("connection_status") == "connected" else "Not Connected"


@router.get("/batches/{batch_id}/report")
async def download_batch_report(batch_id: str, admin: dict = Depends(require_admin)):
    """Excel report for one imported batch: calls made, connection outcomes
    (connected / switched-off / etc.), duplicate & incomplete numbers skipped
    at import time, and a full per-lead breakdown — so admin can audit an
    upload without digging through the UI."""
    import openpyxl
    from fastapi.responses import StreamingResponse

    meta = await lead_batches_collection.find_one({"batch_id": batch_id})
    leads = [doc async for doc in leads_collection.find({"batch_id": batch_id})]
    if not leads and not meta:
        raise HTTPException(status_code=404, detail="Batch not found")

    called = sum(1 for l in leads if l.get("call_touched"))
    status_counts = {}
    for l in leads:
        s = l.get("status", "new")
        status_counts[s] = status_counts.get(s, 0) + 1
    connected = sum(1 for l in leads if l.get("connection_status") == "connected")
    not_connected = sum(1 for l in leads if l.get("connection_status") == "not_connected")

    reason_counts = {}
    for l in leads:
        for h in l.get("status_history", []):
            sub = h.get("sub_stage")
            if sub:
                reason_counts[sub] = reason_counts.get(sub, 0) + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    meta = meta or {}
    ws.append(["Metric", "Value"])
    ws.append(["Batch ID", batch_id])
    ws.append(["Uploaded on", meta.get("created_at", "")])
    ws.append(["Total rows in file", meta.get("total_rows", len(leads))])
    ws.append(["Imported", meta.get("imported_count", len(leads))])
    ws.append(["Incomplete numbers skipped", meta.get("incomplete_count", 0)])
    ws.append(["Duplicate numbers — within this same file", meta.get("duplicate_infile_count", 0)])
    ws.append(["Duplicate numbers — already in database", meta.get("duplicate_existing_count", 0)])
    ws.append(["  ...of which re-assigned instead of skipped", meta.get("reassigned_existing_count", 0)])
    ws.append(["", ""])
    ws.append(["Called", called])
    ws.append(["Not called yet", len(leads) - called])
    ws.append(["Connected", connected])
    ws.append(["Not connected", not_connected])
    ws.append(["", ""])
    ws.append(["Status breakdown", ""])
    for status, count in status_counts.items():
        ws.append([STATUS_LABELS.get(status, status), count])
    ws.append(["", ""])
    ws.append(["Call outcome breakdown (all attempts)", ""])
    for sub, count in reason_counts.items():
        ws.append([SUB_STAGE_LABELS.get(sub, sub), count])

    ws2 = wb.create_sheet("Leads")
    ws2.append(["Name", "Phone", "Assigned To", "Status", "Call Outcome", "Connection Status", "Last Call At", "Call Attempts", "Follow-up Date"])
    for l in leads:
        ws2.append([
            l.get("name"), l.get("phone"), l.get("assigned_to_name") or "Unassigned",
            STATUS_LABELS.get(l.get("status"), l.get("status")), lead_call_outcome_label(l),
            l.get("connection_status") or "", l.get("last_call_at") or "",
            l.get("call_attempts", 0), l.get("follow_up_date") or "",
        ])

    if meta.get("incomplete_samples") or meta.get("duplicate_infile_samples") or meta.get("duplicate_existing_samples"):
        ws3 = wb.create_sheet("Skipped Numbers")
        ws3.append(["Reason", "Name", "Phone (as entered)"])
        for r in meta.get("incomplete_samples", []):
            ws3.append(["Incomplete number", r.get("name"), r.get("phone")])
        for r in meta.get("duplicate_infile_samples", []):
            ws3.append(["Duplicate — within this file", r.get("name"), r.get("phone")])
        for r in meta.get("duplicate_existing_samples", []):
            ws3.append(["Duplicate — already in database", r.get("name"), r.get("phone")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await log_activity(
        admin["sub"], "report_downloaded",
        f"Downloaded the lead report for a batch ({len(leads)} leads)",
        link="/portal/admin/leads",
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="lead_batch_report_{batch_id[:8]}.xlsx"'},
    )


@router.get("/admin-my")
async def admin_my_leads(admin: dict = Depends(require_admin)):
    """Leads the admin assigned to themselves."""
    cursor = leads_collection.find({"assigned_to": admin["sub"]}).sort("updated_at", -1)
    return [to_lead_out(doc) async for doc in cursor]


# NOTE: This catch-all MUST stay below /stats, /my, /website, /career, /batches, /admin-my —
# FastAPI matches routes in registration order, and "/{lead_id}" would
# otherwise swallow those literal paths.
@router.get("/{lead_id}", response_model=LeadOut)
async def get_lead(lead_id: str, payload: dict = Depends(get_current_user_payload)):
    doc = await leads_collection.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    return to_lead_out(doc)


@router.put("/{lead_id}", response_model=LeadOut)
async def update_lead(lead_id: str, data: LeadUpdate, admin: dict = Depends(require_admin)):
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()

    if "assigned_to" in updates:
        emp = await users_collection.find_one({"id": updates["assigned_to"]})
        updates["assigned_to_name"] = emp.get("name") if emp else None
        if "pipeline_id" not in updates:
            updates["pipeline_id"] = await find_pipeline_for_employee(updates["assigned_to"])
        if emp:
            lead_doc = await leads_collection.find_one({"id": lead_id})
            lead_name = lead_doc.get("name", "Unknown") if lead_doc else "Unknown"
            await create_notification(
                user_id=updates["assigned_to"],
                title="Lead Assigned to You",
                body=f"{lead_name} — check your leads dashboard",
                n_type="lead",
                link=f"/portal/employee/leads?leadId={lead_id}",
            )

    result = await leads_collection.find_one_and_update(
        {"id": lead_id},
        {"$set": updates},
        return_document=True,
    )
    if not result:
        raise HTTPException(status_code=404, detail="Lead not found")
    return to_lead_out(result)


@router.delete("/{lead_id}")
async def delete_lead(lead_id: str, admin: dict = Depends(require_admin)):
    result = await leads_collection.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    return {"status": "deleted"}


@router.post("/{lead_id}/assign")
async def assign_lead(lead_id: str, assigned_to: str = Query(...), admin: dict = Depends(require_admin)):
    emp = await users_collection.find_one({"id": assigned_to})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    pipeline_id = await find_pipeline_for_employee(assigned_to)
    result = await leads_collection.find_one_and_update(
        {"id": lead_id},
        {"$set": {
            "assigned_to": assigned_to,
            "assigned_to_name": emp.get("name"),
            "pipeline_id": pipeline_id,
            "pipeline_stage_id": None,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
        return_document=True,
    )
    if not result:
        raise HTTPException(status_code=404, detail="Lead not found")

    await create_notification(
        user_id=assigned_to,
        title="New Lead Assigned",
        body=f"{result['name']} ({result['phone']}) — {result.get('service_interest') or 'General'}",
        n_type="lead",
        link=f"/portal/employee/leads?leadId={lead_id}",
    )
    return {"status": "assigned", "assigned_to_name": emp.get("name")}


# ── Employee: status update ──────────────────────────────────────

@router.post("/{lead_id}/status")
async def update_lead_status(
    lead_id: str, data: LeadStatusUpdate, payload: dict = Depends(get_current_user_payload)
):
    if data.status not in VALID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Use: {', '.join(VALID_STATUSES)}")

    doc = await leads_collection.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")

    if payload["role"] != "admin" and doc.get("assigned_to") != payload["sub"]:
        raise HTTPException(status_code=403, detail="Lead not assigned to you")

    now = datetime.now(timezone.utc).isoformat()
    updater = await users_collection.find_one({"id": payload["sub"]})
    updater_name = updater.get("name", "") if updater else ""
    history_entry = {
        "status": data.status,
        "note": data.follow_up_note,
        "at": now,
        "by": payload["sub"],
        "by_name": updater_name,
    }

    updates = {
        "status": data.status,
        "updated_at": now,
    }
    if data.follow_up_note is not None:
        updates["follow_up_note"] = data.follow_up_note
    if data.follow_up_date is not None:
        updates["follow_up_date"] = data.follow_up_date

    reminder_to_create = None  # (type, due_at, title, body) — mirrors the call-outcome route
    if data.status == "converted":
        if data.service_interest:
            updates["service_interest"] = data.service_interest
        if data.code_name:
            updates["code_name"] = data.code_name
        if data.service_price is not None:
            updates["service_price"] = data.service_price
        if data.service_duration_months:
            updates["service_duration_months"] = data.service_duration_months
            expiry = datetime.now(timezone.utc) + timedelta(days=30 * data.service_duration_months)
            updates["service_expires_at"] = expiry.date().isoformat()
            reminder_due = expiry - timedelta(days=3)
            reminder_to_create = ("service_expiry", reminder_due, "Service Expiring Soon", f"{doc['name']}'s service ends in 3 days — reach out to renew")

    await leads_collection.update_one(
        {"id": lead_id},
        {"$set": updates, "$push": {"status_history": history_entry}},
    )

    if reminder_to_create:
        rtype, due_dt, title, body = reminder_to_create
        await reminders_collection.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": doc.get("assigned_to") or payload["sub"],
            "lead_id": lead_id,
            "type": rtype,
            "title": title,
            "body": body,
            "next_send_at": due_dt,
            "active": True,
            "created_at": now,
        })

    # Notify admin when employee updates lead status
    if payload["role"] != "admin":
        user = await users_collection.find_one({"id": payload["sub"]})
        emp_name = user.get("name", "Employee") if user else "Employee"
        admins = users_collection.find({"role": "admin"})
        async for adm in admins:
            await create_notification(
                user_id=adm["id"],
                title=f"Lead Updated: {doc['name']}",
                body=f"{emp_name} changed status to '{data.status}'",
                n_type="lead",
                link=f"/portal/admin/leads?leadId={lead_id}",
            )

    return {"status": "updated", "new_status": data.status}


@router.put("/{lead_id}/name")
async def update_lead_name(
    lead_id: str, data: LeadNameUpdate, payload: dict = Depends(get_current_user_payload)
):
    """Lets whoever owns a lead fill in a name when only a phone number was
    captured (e.g. a missed-call/website lead) — admin's PUT /{lead_id} is
    admin-only, so employees need this narrower endpoint."""
    name = data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name cannot be empty")

    doc = await leads_collection.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    if payload["role"] != "admin" and doc.get("assigned_to") != payload["sub"]:
        raise HTTPException(status_code=403, detail="Lead not assigned to you")

    now = datetime.now(timezone.utc).isoformat()
    await leads_collection.update_one({"id": lead_id}, {"$set": {"name": name, "updated_at": now}})
    return {"status": "updated", "name": name}


@router.put("/{lead_id}/alternate-phone")
async def update_lead_alternate_phone(
    lead_id: str, data: LeadAlternatePhoneUpdate, payload: dict = Depends(get_current_user_payload)
):
    """A second number for the same client (e.g. they call back from a
    different phone) — kept on this one lead rather than creating a
    duplicate lead, mirrors update_lead_name's employee-accessible pattern."""
    alt = data.alternate_phone.strip()
    if not alt:
        raise HTTPException(status_code=400, detail="Alternate phone cannot be empty")

    doc = await leads_collection.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    if payload["role"] != "admin" and doc.get("assigned_to") != payload["sub"]:
        raise HTTPException(status_code=403, detail="Lead not assigned to you")

    now = datetime.now(timezone.utc).isoformat()
    await leads_collection.update_one({"id": lead_id}, {"$set": {"alternate_phone": alt, "updated_at": now}})
    return {"status": "updated", "alternate_phone": alt}


class QuickNoteIn(BaseModel):
    note: str


@router.post("/{lead_id}/note")
async def add_quick_note(
    lead_id: str, data: QuickNoteIn, payload: dict = Depends(get_current_user_payload)
):
    """Add a manual update/note to a lead any time — not tied to a call.
    Used by the 'Add Update' button on both the employee and admin lead views."""
    doc = await leads_collection.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    if payload["role"] != "admin" and doc.get("assigned_to") != payload["sub"]:
        raise HTTPException(status_code=403, detail="Lead not assigned to you")

    now = datetime.now(timezone.utc).isoformat()
    who = await users_collection.find_one({"id": payload["sub"]})
    history_entry = {
        "note": data.note,
        "date": now,
        "by": payload["sub"],
        "by_name": who.get("name") if who else "Unknown",
    }
    await leads_collection.update_one(
        {"id": lead_id},
        {"$set": {"notes": data.note, "updated_at": now}, "$push": {"status_history": history_entry}},
    )
    return {"status": "noted"}


@router.post("/{lead_id}/call-started")
async def mark_call_started(lead_id: str, payload: dict = Depends(get_current_user_payload)):
    """Fired the moment the employee taps the tel: link — before the call
    happens or any outcome is known. Lets a "Recent Calls" widget recover a
    call whose outcome popup was missed/dismissed (app backgrounded and
    killed mid-call, accidental tap-away, etc.), since last_call_at only
    gets set once an outcome is actually submitted."""
    doc = await leads_collection.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    if payload["role"] != "admin" and doc.get("assigned_to") != payload["sub"]:
        raise HTTPException(status_code=403, detail="Lead not assigned to you")
    now = datetime.now(timezone.utc).isoformat()
    await leads_collection.update_one({"id": lead_id}, {"$set": {"last_call_attempted_at": now}})
    return {"status": "ok", "last_call_attempted_at": now}


# ── Employee: full call-flow outcome (the popup after every call) ──

CONNECTED_TERMINAL = {"converted", "lost"}
NOT_CONNECTED_DIRECT_LOST = {"invalid"}  # invalid number -> lost immediately, no retries
MAX_NOT_INTERESTED_ATTEMPTS = 3
MAX_REASSIGNS = 2  # after this many hand-offs with still no connect, give up and mark lost


async def _pick_reassign_target(exclude_id: Optional[str]) -> Optional[dict]:
    """Pick another active employee to hand a stalled lead off to — the one
    with the fewest currently-open leads, so hand-offs don't all pile onto
    whichever employee happens to be queried first."""
    candidates = [
        e async for e in users_collection.find({"role": "employee", "is_active": {"$ne": False}})
        if e["id"] != exclude_id
    ]
    best, best_load = None, None
    for emp in candidates:
        load = await leads_collection.count_documents({
            "assigned_to": emp["id"], "status": {"$nin": ["lost", "converted"]},
        })
        if best_load is None or load < best_load:
            best, best_load = emp, load
    return best


@router.post("/{lead_id}/call-outcome")
async def submit_call_outcome(
    lead_id: str, data: CallOutcomeIn, payload: dict = Depends(get_current_user_payload)
):
    """Saves the result of the Connected/Not-Connected decision tree shown
    right after an employee finishes a call, and drives what happens next:
    reminders, auto-lost, auto-reassignment, and service-expiry reminders.

    Not-connected rules:
      - invalid number -> lost immediately, no retry.
      - switched off / network issue / busy / NPC -> retry same employee next
        day; a 2nd consecutive miss hands the lead to a different employee
        (up to 2 hand-offs total) before it's finally marked lost with the
        full reason trail in status_history/transfer_history.

    Connected + Not Interested -> normal retry-then-lost flow, unless the
    employee opts (via `reassign_to`) to hand it straight to another
    employee, carrying the notes/service_interest forward.
    """
    doc = await leads_collection.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    if payload["role"] != "admin" and doc.get("assigned_to") != payload["sub"]:
        raise HTTPException(status_code=403, detail="Lead not assigned to you")
    if doc.get("status") == "lost":
        # Calling an already-lost lead through this flow again (easy to do
        # by mistake, repeatedly) would otherwise write a fresh history
        # entry and re-fire notifications every single time. If they really
        # need to change it, that's a direct status edit elsewhere, not
        # another "call outcome" submission.
        raise HTTPException(status_code=400, detail="This lead is already marked Lost. Use the status editor if you need to change it to something else.")

    now = datetime.now(timezone.utc)
    now_iso = now.isoformat()
    updates = {
        "connection_status": data.connection_status,
        "last_call_at": now_iso,
        "updated_at": now_iso,
        "call_touched": True,
    }
    if data.pipeline_stage_id:
        updates["pipeline_stage_id"] = data.pipeline_stage_id
    history_entry = {"connection_status": data.connection_status, "sub_stage": data.sub_stage, "date": now_iso, "by": payload["sub"]}
    if data.notes:
        history_entry["note"] = data.notes

    reminder_to_create = None  # (type, due_at, title, body)
    transfer_entry = None      # set when the lead is handed to a different employee
    notify_new_assignee = None # employee id to immediately notify of a hand-off

    if data.connection_status == "connected":
        if data.sub_stage == "interested":
            updates["status"] = "interested"
            updates["call_attempts"] = 0
            if data.service_interest:
                updates["service_interest"] = data.service_interest
            if data.notes:
                updates["notes"] = data.notes
            if data.follow_up_date:
                updates["follow_up_date"] = data.follow_up_date
                due_dt = _combine_date_time(data.follow_up_date, data.follow_up_time)
                reminder_to_create = ("lead_follow_up", due_dt, "Follow-up Due", f"Time to follow up with {doc['name']} ({doc['phone']})")

        elif data.sub_stage == "follow_up":
            # A direct "schedule a follow-up call" outcome, distinct from
            # "Interested" — the call connected but the employee just wants
            # to check back later without committing to an interest level yet.
            updates["status"] = "follow_up"
            updates["call_attempts"] = 0
            if data.notes:
                updates["notes"] = data.notes
            if data.follow_up_date:
                updates["follow_up_date"] = data.follow_up_date
                due_dt = _combine_date_time(data.follow_up_date, data.follow_up_time)
                reminder_to_create = ("lead_follow_up", due_dt, "Follow-up Due", f"Time to follow up with {doc['name']} ({doc['phone']})")

        elif data.sub_stage == "not_interested":
            if data.reassign_to:
                new_emp = await users_collection.find_one({"id": data.reassign_to})
                if not new_emp:
                    raise HTTPException(status_code=404, detail="Employee to reassign to not found")
                transfer_entry = {
                    "from": doc.get("assigned_to"), "from_name": doc.get("assigned_to_name"),
                    "to": data.reassign_to, "to_name": new_emp.get("name"),
                    "note": f"Reassigned after 'Not Interested'" + (f" — {data.notes}" if data.notes else ""),
                    "at": now_iso,
                }
                updates["assigned_to"] = data.reassign_to
                updates["assigned_to_name"] = new_emp.get("name")
                updates["call_attempts"] = 0
                updates["status"] = "new"
                if data.notes:
                    updates["notes"] = data.notes
                notify_new_assignee = data.reassign_to
            else:
                new_attempts = int(doc.get("call_attempts", 0)) + 1
                updates["call_attempts"] = new_attempts
                if new_attempts >= MAX_NOT_INTERESTED_ATTEMPTS:
                    updates["status"] = "lost"
                else:
                    updates["status"] = "follow_up"
                    due_dt = now + timedelta(days=1)
                    reminder_to_create = ("lead_retry", due_dt, "Retry Call", f"{doc['name']} wasn't interested last time — try again today")

        elif data.sub_stage == "converted":
            updates["status"] = "converted"
            updates["call_attempts"] = 0
            if data.service_interest:
                updates["service_interest"] = data.service_interest
            if data.code_name:
                updates["code_name"] = data.code_name
            if data.service_price is not None:
                updates["service_price"] = data.service_price
            if data.service_duration_months:
                updates["service_duration_months"] = data.service_duration_months
                expiry = now + timedelta(days=30 * data.service_duration_months)
                updates["service_expires_at"] = expiry.date().isoformat()
                reminder_due = expiry - timedelta(days=3)
                reminder_to_create = ("service_expiry", reminder_due, "Service Expiring Soon", f"{doc['name']}'s service ends in 3 days — reach out to renew")

        elif data.sub_stage == "lost":
            updates["status"] = "lost"

    else:  # not_connected
        reason_label = REASON_LABELS.get(data.sub_stage, data.sub_stage)
        if data.sub_stage in NOT_CONNECTED_DIRECT_LOST:
            updates["status"] = "lost"
            updates["call_attempts"] = 0
            history_entry["note"] = f"Lost — {reason_label}, no retry for invalid numbers."
        else:
            new_attempts = int(doc.get("call_attempts", 0)) + 1
            updates["call_attempts"] = new_attempts
            # Once real contact has ever happened on this lead (a prior
            # "connected" outcome, at any point in its history), the
            # auto-reassign-after-2-misses safety net no longer applies —
            # that net exists only to rescue leads that were NEVER reached
            # by the current employee. If the employee already has an
            # ongoing conversation with this client, missing a follow-up
            # call any number of times should just keep retrying with the
            # same employee, never silently hand the client to someone else.
            ever_connected = any(h.get("connection_status") == "connected" for h in doc.get("status_history", []))
            if new_attempts == 1 or ever_connected:
                updates["status"] = "follow_up"
                due_dt = now + timedelta(days=1)
                reminder_to_create = ("lead_retry", due_dt, "Retry Call", f"Couldn't reach {doc['name']} ({reason_label}) — try again today")
            else:
                reassign_count = int(doc.get("reassign_count", 0))
                if reassign_count >= MAX_REASSIGNS:
                    updates["status"] = "lost"
                    updates["call_attempts"] = 0
                    history_entry["note"] = f"Lost — no response after {reassign_count} reassignment(s) across employees. Last reason: {reason_label}."
                else:
                    new_emp = await _pick_reassign_target(doc.get("assigned_to"))
                    if not new_emp:
                        updates["status"] = "lost"
                        updates["call_attempts"] = 0
                        history_entry["note"] = f"Lost — no response ({reason_label}), no other active employee available to reassign to."
                    else:
                        transfer_entry = {
                            "from": doc.get("assigned_to"), "from_name": doc.get("assigned_to_name"),
                            "to": new_emp["id"], "to_name": new_emp.get("name"),
                            "note": f"Auto-reassigned — 2 consecutive not-connected attempts ({reason_label})",
                            "at": now_iso,
                        }
                        updates["assigned_to"] = new_emp["id"]
                        updates["assigned_to_name"] = new_emp.get("name")
                        updates["call_attempts"] = 0
                        updates["reassign_count"] = reassign_count + 1
                        updates["status"] = "follow_up"
                        notify_new_assignee = new_emp["id"]

    update_op = {"$set": updates, "$push": {"status_history": history_entry}}
    if transfer_entry:
        update_op["$push"]["transfer_history"] = transfer_entry
    await leads_collection.update_one({"id": lead_id}, update_op)

    if reminder_to_create:
        rtype, due_dt, title, body = reminder_to_create
        await reminders_collection.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": updates.get("assigned_to") or doc.get("assigned_to") or payload["sub"],
            "lead_id": lead_id,
            "type": rtype,
            "title": title,
            "body": body,
            "next_send_at": due_dt if isinstance(due_dt, datetime) else due_dt,
            "active": True,
            "created_at": now_iso,
        })

    if notify_new_assignee:
        await create_notification(
            user_id=notify_new_assignee,
            title="Lead Reassigned to You",
            body=f"{doc['name']} ({doc['phone']}) — {transfer_entry['note'] if transfer_entry else ''}",
            n_type="lead",
            link=f"/portal/employee/leads?leadId={lead_id}",
        )

    # If the lead just hit final auto-lost, flag admin so they know it happened
    if updates.get("status") == "lost":
        admins = users_collection.find({"role": "admin"})
        async for adm in admins:
            await create_notification(
                user_id=adm["id"],
                title=f"Lead marked Lost: {doc['name']}",
                body=history_entry.get("note") or "No result after repeated attempts.",
                n_type="lead",
                link=f"/portal/admin/leads?leadId={lead_id}",
            )

    return {"status": "saved", "new_status": updates.get("status"), "call_attempts": updates.get("call_attempts", doc.get("call_attempts", 0))}


def _combine_date_time(date_str: str, time_str: Optional[str]) -> datetime:
    t = time_str or "09:00"
    try:
        return datetime.fromisoformat(f"{date_str}T{t}:00")
    except Exception:
        return datetime.now(timezone.utc) + timedelta(days=1)


# ── Transfer a lead to another employee ─────────────────────────

@router.post("/{lead_id}/transfer")
async def transfer_lead(
    lead_id: str, data: TransferIn, payload: dict = Depends(get_current_user_payload)
):
    doc = await leads_collection.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    if payload["role"] != "admin" and doc.get("assigned_to") != payload["sub"]:
        raise HTTPException(status_code=403, detail="Lead not assigned to you")

    new_emp = await users_collection.find_one({"id": data.to_employee_id})
    if not new_emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    now_iso = datetime.now(timezone.utc).isoformat()
    transfer_entry = {
        "from": doc.get("assigned_to"),
        "from_name": doc.get("assigned_to_name"),
        "to": data.to_employee_id,
        "to_name": new_emp.get("name"),
        "note": data.reference_note,
        "at": now_iso,
    }

    await leads_collection.update_one(
        {"id": lead_id},
        {
            "$set": {
                "assigned_to": data.to_employee_id,
                "assigned_to_name": new_emp.get("name"),
                "call_attempts": 0,
                "updated_at": now_iso,
            },
            "$push": {"transfer_history": transfer_entry},
        },
    )

    await create_notification(
        user_id=data.to_employee_id,
        title="Lead Transferred to You",
        body=f"{doc['name']} ({doc['phone']}) — {data.reference_note or 'no note'}",
        n_type="lead",
        link=f"/portal/employee/leads?leadId={lead_id}",
    )
    return {"status": "transferred", "to_name": new_emp.get("name")}


# ── Employee: add a lead themselves ─────────────────────────────

@router.post("/my", response_model=LeadOut)
async def create_my_lead(
    data: LeadCreate,
    confirm_merge: bool = Query(default=False),
    payload: dict = Depends(get_current_user_payload),
):
    """Any employee can add a lead directly (referrals, walk-ins, etc.) — it's
    auto-assigned to them. Always tagged source="employee_added" (regardless
    of whatever LeadCreate.source default came through) so both the employee
    and admin can tell it was self-added rather than assigned from elsewhere.
    If referred_by_lead_id is set, this becomes a structural referral: source
    flips to "referral" and reference_note is auto-filled from the referring
    client's name (unless already supplied).

    Phone must be a complete, valid 10-digit Indian mobile number (normalized
    the same way import-excel dedupes — see normalize_phone/is_complete_phone
    below). If the number already exists among this employee's own leads,
    the create is rejected with 409 + the existing lead's summary unless the
    caller passes confirm_merge=true, in which case the new submission's
    details are folded into the existing lead (as a note) instead of creating
    a duplicate record."""
    norm_new = normalize_phone(data.phone)
    if not is_complete_phone(norm_new) or not re.match(r"^[6-9]\d{9}$", norm_new):
        raise HTTPException(
            status_code=400,
            detail="Enter a valid 10-digit Indian mobile number (must start with 6-9).",
        )

    existing = None
    async for doc in leads_collection.find(
        {"assigned_to": payload["sub"]},
        {"id": 1, "name": 1, "phone": 1, "status": 1, "service_interest": 1, "created_at": 1},
    ):
        if normalize_phone(doc.get("phone")) == norm_new:
            existing = doc
            break

    if existing and not confirm_merge:
        raise HTTPException(
            status_code=409,
            detail={
                "message": f"{data.name.strip() or 'This number'} ({data.phone}) is already in your leads as \"{existing.get('name')}\".",
                "existing_lead": {
                    "id": existing["id"],
                    "name": existing.get("name"),
                    "phone": existing.get("phone"),
                    "status": existing.get("status"),
                    "service_interest": existing.get("service_interest"),
                    "created_at": existing.get("created_at"),
                },
            },
        )

    if existing and confirm_merge:
        now = datetime.now(timezone.utc).isoformat()
        who = await users_collection.find_one({"id": payload["sub"]})
        note_bits = [f"Merged a duplicate re-add (name entered: \"{data.name.strip()}\")"]
        update: dict = {"updated_at": now}
        if data.service_interest and data.service_interest != existing.get("service_interest"):
            note_bits.append(f"service interest: {data.service_interest}")
            update["service_interest"] = data.service_interest
        if data.notes:
            note_bits.append(f"note: {data.notes}")
        if data.city:
            update["city"] = data.city
        if data.email:
            update["email"] = data.email
        history_entry = {
            "note": " — ".join(note_bits),
            "date": now,
            "by": payload["sub"],
            "by_name": who.get("name") if who else "Unknown",
        }
        await leads_collection.update_one(
            {"id": existing["id"]},
            {"$set": update, "$push": {"status_history": history_entry}},
        )
        merged = await leads_collection.find_one({"id": existing["id"]})
        return to_lead_out(merged)

    user = await users_collection.find_one({"id": payload["sub"]})
    emp_name = user.get("name") if user else None
    lead = LeadInDB(**data.model_dump())
    lead.assigned_to = payload["sub"]
    lead.assigned_to_name = emp_name
    lead.source = "employee_added"

    if data.referred_by_lead_id:
        referrer = await leads_collection.find_one({"id": data.referred_by_lead_id})
        if referrer:
            lead.source = "referral"
            if not lead.reference_note:
                lead.reference_note = f"Referred by {referrer.get('name', 'a client')} ({referrer.get('phone', '')})"
        else:
            lead.referred_by_lead_id = None

    lead.pipeline_id = await find_pipeline_for_employee(payload["sub"])
    await leads_collection.insert_one(lead.model_dump())

    await notify_admins_employee_added_lead(payload["sub"], emp_name or "An employee", lead.name, lead.phone)

    return to_lead_out(lead.model_dump())


# ── Employee: update service_interest on their own lead ──────────

class ServiceUpdateBody(BaseModel):
    service_interest: str


@router.put("/{lead_id}/service")
async def update_my_lead_service(lead_id: str, data: ServiceUpdateBody, payload: dict = Depends(get_current_user_payload)):
    """Employee can add a service to their assigned lead."""
    doc = await leads_collection.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    if payload["role"] != "admin" and doc.get("assigned_to") != payload["sub"]:
        raise HTTPException(status_code=403, detail="Lead not assigned to you")
    now = datetime.now(timezone.utc).isoformat()
    await leads_collection.update_one(
        {"id": lead_id},
        {"$set": {"service_interest": data.service_interest, "updated_at": now}},
    )
    return {"status": "updated"}


# ── Excel Import ─────────────────────────────────────────────────

def normalize_phone(raw: Optional[str]) -> str:
    """Strips everything but digits and drops a leading country code (91/0)
    so the same number typed as '+91 98765 43210', '09876543210' or
    '9876543210' all dedupe to the same 10-digit key."""
    digits = re.sub(r"\D", "", raw or "")
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    return digits


def is_complete_phone(digits: str) -> bool:
    return len(digits) == 10


def _auto_detect_columns(headers_lower: list[str]) -> dict:
    col_map = {}
    for i, h in enumerate(headers_lower):
        if "name" in h and "name" not in col_map:
            col_map["name"] = i
        elif "phone" in h or "contact" in h or "mobile" in h:
            col_map["phone"] = i
        elif "mail" in h or "email" in h:
            col_map["email"] = i
        elif "from" in h or "source" in h:
            col_map["source"] = i
        elif "for" in h or "service" in h or "interest" in h:
            col_map["service_interest"] = i
        elif "city" in h:
            col_map["city"] = i
    return col_map


@router.post("/import-excel/preview")
async def preview_leads_excel(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    """Returns the header row, a few sample data rows, and our best guess at
    which column is which — the admin UI shows this so a column can be
    re-mapped by hand before the real import runs, instead of the import
    just failing outright when headers don't match our guesses (e.g. an
    Excel with only a phone-number column, or headers in a different
    wording/language)."""
    import openpyxl
    data = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not all_rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    headers_row = list(all_rows[0])
    headers = [str(h).strip() if h is not None else "" for h in headers_row]
    headers_lower = [h.lower() for h in headers]
    auto_map = _auto_detect_columns(headers_lower)
    sample_rows = [[("" if c is None else str(c)) for c in row] for row in all_rows[1:6]]

    return {
        "headers": headers,
        "sample_rows": sample_rows,
        "total_rows": len(all_rows) - 1,
        "auto_map": auto_map,
    }


@router.post("/import-excel")
async def import_leads_excel(
    file: UploadFile = File(...),
    assign_to: Optional[str] = Query(default=None),
    assign_mode: Optional[str] = Query(default=None),  # "self", "all", "selected" — required, no unassigned imports
    employee_ids: Optional[str] = Query(default=None),  # comma-separated
    include_existing_duplicates: bool = Query(default=False),  # if True, numbers already in the DB get re-assigned instead of skipped
    name_col: Optional[int] = Query(default=None),  # explicit column-mapping override (from the preview step) — -1 means "not present"
    phone_col: Optional[int] = Query(default=None),
    email_col: Optional[int] = Query(default=None),
    source_col: Optional[int] = Query(default=None),
    service_interest_col: Optional[int] = Query(default=None),
    city_col: Optional[int] = Query(default=None),
    admin: dict = Depends(require_admin),
):
    import openpyxl
    data = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers_lower = [str(h).strip().lower() if h else "" for h in headers_row]

    # An explicit mapping from the preview/confirm step always wins; anything
    # left unset (None, not passed at all) falls back to auto-detection so
    # existing callers that never used the preview step keep working.
    explicit = {
        "name": name_col, "phone": phone_col, "email": email_col,
        "source": source_col, "service_interest": service_interest_col, "city": city_col,
    }
    col_map = _auto_detect_columns(headers_lower)
    for key, idx in explicit.items():
        if idx is None:
            continue
        if idx < 0:
            col_map.pop(key, None)  # explicitly marked "not present" — don't fall back to a guess
        else:
            col_map[key] = idx

    # A name column is no longer required — a lead with just a phone number
    # is still a real lead (e.g. a missed call), and can have its name filled
    # in later from the employee portal (PUT /{lead_id}/name).
    if "phone" not in col_map:
        raise HTTPException(status_code=400, detail="Couldn't find a Phone/Contact column — please map one in the import screen.")

    # Determine assignment targets — importing without assigning to someone
    # (self or one/more employees) is no longer allowed.
    target_employees = []
    if assign_mode == "self":
        target_employees = [admin["sub"]]
    elif assign_mode == "all":
        all_emps = users_collection.find({"role": "employee", "is_active": {"$ne": False}})
        target_employees = [emp["id"] async for emp in all_emps]
    elif assign_mode == "selected" and employee_ids:
        target_employees = [eid.strip() for eid in employee_ids.split(",") if eid.strip()]
    elif assign_to:
        target_employees = [assign_to]
    if not target_employees:
        raise HTTPException(
            status_code=400,
            detail="Choose who to assign these leads to — yourself, all employees, or specific employees.",
        )

    # Resolve names and pipelines for targets
    emp_names = {}
    emp_pipelines = {}
    for eid in target_employees:
        emp = await users_collection.find_one({"id": eid})
        if emp:
            emp_names[eid] = emp.get("name", "")
        emp_pipelines[eid] = await find_pipeline_for_employee(eid)

    # Every phone already in the system, normalized, so a duplicate is caught
    # regardless of how it was formatted the first time it was entered.
    # existing_phone_to_id lets an "include existing duplicates" import
    # re-assign the lead that's already there instead of creating a second,
    # literally-duplicate lead document for the same number.
    existing_phones = set()
    existing_phone_to_id = {}
    async for doc in leads_collection.find({}, {"phone": 1, "id": 1}):
        norm = normalize_phone(doc.get("phone"))
        if norm:
            existing_phones.add(norm)
            existing_phone_to_id.setdefault(norm, doc.get("id"))
    seen_in_file = set()

    batch_id = str(uuid.uuid4())
    now_iso = datetime.now(timezone.utc).isoformat()
    imported = 0
    reassigned_existing = 0
    imported_ids = []
    incomplete_rows = []          # [{name, phone}] — not a complete 10-digit number
    duplicate_infile_rows = []    # [{name, phone}] — repeated more than once within this same upload
    duplicate_existing_rows = []  # [{name, phone}] — already present in the database from before

    def _cell(row, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    for row in rows:
        name = _cell(row, "name")
        phone = _cell(row, "phone")
        if not phone:
            continue

        norm_phone = normalize_phone(phone)
        if not is_complete_phone(norm_phone):
            incomplete_rows.append({"name": name, "phone": phone})
            continue
        if norm_phone in seen_in_file:
            duplicate_infile_rows.append({"name": name, "phone": phone})
            continue
        seen_in_file.add(norm_phone)

        if norm_phone in existing_phones:
            duplicate_existing_rows.append({"name": name, "phone": phone})
            if not include_existing_duplicates:
                continue
            # Re-assign the existing lead to the next target rather than
            # inserting a second lead document for the same phone number.
            existing_id = existing_phone_to_id.get(norm_phone)
            if existing_id:
                eid = target_employees[imported % len(target_employees)]
                await leads_collection.update_one(
                    {"id": existing_id},
                    {"$set": {
                        "assigned_to": eid, "assigned_to_name": emp_names.get(eid, ""),
                        "pipeline_id": emp_pipelines.get(eid), "pipeline_stage_id": None,
                        "batch_id": batch_id, "batch_date": now_iso,
                        "updated_at": now_iso,
                    }},
                )
                imported += 1
                reassigned_existing += 1
                imported_ids.append(existing_id)
            continue

        lead = LeadInDB(
            name=name,
            phone=phone,
            email=_cell(row, "email") or None,
            source=_cell(row, "source") or "excel",
            service_interest=_cell(row, "service_interest") or None,
            city=_cell(row, "city") or None,
        )
        lead_doc = lead.model_dump()
        lead_doc["batch_id"] = batch_id
        lead_doc["batch_date"] = now_iso

        eid = target_employees[imported % len(target_employees)]
        lead_doc["assigned_to"] = eid
        lead_doc["assigned_to_name"] = emp_names.get(eid, "")
        lead_doc["pipeline_id"] = emp_pipelines.get(eid)

        await leads_collection.insert_one(lead_doc)
        imported += 1
        imported_ids.append(lead.id)
    wb.close()

    await lead_batches_collection.insert_one({
        "batch_id": batch_id,
        "total_rows": len(rows),
        "imported_count": imported,
        "reassigned_existing_count": reassigned_existing,
        "incomplete_count": len(incomplete_rows),
        "duplicate_infile_count": len(duplicate_infile_rows),
        "duplicate_existing_count": len(duplicate_existing_rows),
        "incomplete_samples": incomplete_rows[:50],
        "duplicate_infile_samples": duplicate_infile_rows[:50],
        "duplicate_existing_samples": duplicate_existing_rows[:50],
        "included_existing_duplicates": include_existing_duplicates,
        "created_at": now_iso,
    })

    # Send notifications to assigned employees
    assignment_counts = {}
    for i, lid in enumerate(imported_ids):
        eid = target_employees[i % len(target_employees)]
        assignment_counts[eid] = assignment_counts.get(eid, 0) + 1
    for eid, count in assignment_counts.items():
        await create_notification(
            user_id=eid,
            title=f"{count} New Leads Assigned",
            body=f"{count} leads from Excel import have been assigned to you",
            n_type="lead",
            link="/portal/employee/leads",
        )

    await log_activity(
        admin["sub"], "lead_imported",
        f"Imported {imported} leads via Excel ({len(incomplete_rows)} incomplete, "
        f"{len(duplicate_infile_rows)} duplicate-in-file, {len(duplicate_existing_rows)} duplicate-in-database"
        f"{f', {reassigned_existing} of those re-assigned' if reassigned_existing else ''})",
        link="/portal/admin/leads",
    )

    return {
        "status": "imported",
        "count": imported,
        "reassigned_existing_count": reassigned_existing,
        "lead_ids": imported_ids,
        "batch_id": batch_id,
        "incomplete_count": len(incomplete_rows),
        "duplicate_infile_count": len(duplicate_infile_rows),
        "duplicate_existing_count": len(duplicate_existing_rows),
    }


# ── Shuffle & Assign leads to employees ──────────────────────────

@router.post("/shuffle-assign")
async def shuffle_assign_leads(request: Request, admin: dict = Depends(require_admin)):
    body = await request.json()
    lead_ids = body.get("lead_ids", [])
    employee_ids = body.get("employee_ids", [])
    if not lead_ids or not employee_ids:
        raise HTTPException(status_code=400, detail="Provide lead_ids and employee_ids")
    random.shuffle(lead_ids)
    emp_names = {}
    emp_pipelines = {}
    for eid in employee_ids:
        emp = await users_collection.find_one({"id": eid})
        if emp:
            emp_names[eid] = emp.get("name", "")
        emp_pipelines[eid] = await find_pipeline_for_employee(eid)
    assignments = []
    for i, lid in enumerate(lead_ids):
        eid = employee_ids[i % len(employee_ids)]
        await leads_collection.update_one(
            {"id": lid},
            {"$set": {
                "assigned_to": eid,
                "assigned_to_name": emp_names.get(eid, ""),
                "pipeline_id": emp_pipelines.get(eid),
                "pipeline_stage_id": None,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }},
        )
        assignments.append({"lead_id": lid, "employee_id": eid, "employee_name": emp_names.get(eid, "")})
    for eid in employee_ids:
        count = sum(1 for a in assignments if a["employee_id"] == eid)
        if count > 0:
            await create_notification(
                user_id=eid,
                title=f"{count} Leads Assigned",
                body=f"{count} new leads have been assigned to you",
                n_type="lead",
                link="/portal/employee/leads",
            )
    return {"status": "assigned", "total": len(assignments), "assignments": assignments}


# ── Website Leads (from contact form/popup) ──────────────────────


@router.post("/website/{web_lead_id}/convert")
async def convert_web_lead(
    web_lead_id: str,
    assign_to: Optional[str] = Query(default=None),
    admin: dict = Depends(require_admin),
):
    wl = await web_leads_collection.find_one({"id": web_lead_id})
    if not wl:
        raise HTTPException(status_code=404, detail="Website lead not found")
    lead = LeadInDB(
        name=wl.get("full_name") or wl.get("name", "Unknown"),
        phone=wl.get("phone", ""),
        email=wl.get("email"),
        city=wl.get("city"),
        source="website",
        service_interest=wl.get("service"),
        notes=wl.get("message"),
    )
    if assign_to:
        emp = await users_collection.find_one({"id": assign_to})
        if emp:
            lead.assigned_to = assign_to
            lead.assigned_to_name = emp.get("name")
            lead.pipeline_id = await find_pipeline_for_employee(assign_to)
    await leads_collection.insert_one(lead.model_dump())
    await web_leads_collection.update_one({"id": web_lead_id}, {"$set": {"converted": True}})
    if lead.assigned_to:
        await create_notification(
            user_id=lead.assigned_to,
            title="New Lead Assigned",
            body=f"{lead.name} ({lead.phone}) — {lead.service_interest or 'Website enquiry'}",
            n_type="lead",
            link=f"/portal/employee/leads?leadId={lead.id}",
        )
    return {"status": "converted", "lead_id": lead.id}


# ── Career Leads (from Career page applications) ──────────────────

@router.post("/career/{career_lead_id}/convert")
async def convert_career_lead(
    career_lead_id: str,
    assign_to: Optional[str] = Query(default=None),
    admin: dict = Depends(require_admin),
):
    """Turn a career application into a callable lead (e.g. for follow-up
    or referring the applicant to a paid service), optionally assigning it
    to an employee in the same step."""
    cl = await career_leads_collection.find_one({"id": career_lead_id})
    if not cl:
        raise HTTPException(status_code=404, detail="Career lead not found")
    lead = LeadInDB(
        name=cl.get("full_name", "Unknown"),
        phone=cl.get("phone", ""),
        email=cl.get("email"),
        source="career",
        service_interest=cl.get("position"),
        notes=cl.get("message"),
    )
    if assign_to:
        emp = await users_collection.find_one({"id": assign_to})
        if emp:
            lead.assigned_to = assign_to
            lead.assigned_to_name = emp.get("name")
            lead.pipeline_id = await find_pipeline_for_employee(assign_to)
    await leads_collection.insert_one(lead.model_dump())
    await career_leads_collection.update_one({"id": career_lead_id}, {"$set": {"converted": True}})
    if lead.assigned_to:
        await create_notification(
            user_id=lead.assigned_to,
            title="New Lead Assigned",
            body=f"{lead.name} ({lead.phone}) — Career applicant: {lead.service_interest or 'General'}",
            n_type="lead",
            link=f"/portal/employee/leads?leadId={lead.id}",
        )
    return {"status": "converted", "lead_id": lead.id}


# ── Admin: delete an entire imported batch ────────────────────────

@router.delete("/batches/{batch_id}")
async def delete_batch(batch_id: str, admin: dict = Depends(require_admin)):
    result = await leads_collection.delete_many({"batch_id": batch_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Batch not found")
    return {"status": "deleted", "count": result.deleted_count}


@router.delete("/website/{web_lead_id}")
async def delete_web_lead(web_lead_id: str, admin: dict = Depends(require_admin)):
    result = await web_leads_collection.delete_one({"id": web_lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"status": "deleted"}


# ── Public capture endpoints (NO auth) — called directly from the public website ──
# Used by Contact form / popups (website leads) and the Career page (career leads).

class PublicWebsiteLeadCreate(BaseModel):
    full_name: str = Field(min_length=2, max_length=100)
    phone: str = Field(min_length=8, max_length=15)
    email: Optional[str] = None
    city: Optional[str] = None
    service: Optional[str] = None
    message: Optional[str] = None
    source: Optional[str] = "website"  # website, popup, calculator, whatsapp_popup, etc.


class PublicCareerLeadCreate(BaseModel):
    full_name: str = Field(min_length=2, max_length=100)
    phone: str = Field(min_length=8, max_length=15)
    email: Optional[str] = None
    position: Optional[str] = None
    experience: Optional[str] = None
    message: Optional[str] = None
    resume_url: Optional[str] = None
    # Extended application fields (personal, contact, position, education, skills, docs)
    father_name: Optional[str] = None
    mother_name: Optional[str] = None
    dob: Optional[str] = None
    gender: Optional[str] = None
    marital_status: Optional[str] = None
    alt_mobile: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    address: Optional[str] = None
    current_company: Optional[str] = None
    current_designation: Optional[str] = None
    current_salary: Optional[str] = None
    expected_salary: Optional[str] = None
    qualification: Optional[str] = None
    college: Optional[str] = None
    passing_year: Optional[str] = None
    skills: Optional[List[str]] = None
    other_skills: Optional[str] = None
    ref_name: Optional[str] = None
    why_join: Optional[str] = None
    additional_info: Optional[str] = None
    photo_url: Optional[str] = None
    hear_about_us: Optional[str] = None
    hear_about_us_other: Optional[str] = None


@router.post("/public/website")
async def capture_website_lead(data: PublicWebsiteLeadCreate):
    """Public, unauthenticated endpoint. Any website contact form / popup should
    POST here so the lead shows up under Admin Portal -> Leads -> Web Leads."""
    doc = {
        "id": str(uuid.uuid4()),
        "full_name": data.full_name,
        "phone": data.phone,
        "email": data.email,
        "city": data.city,
        "service": data.service,
        "message": data.message,
        "source": data.source or "website",
        "converted": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await web_leads_collection.insert_one(doc)
    doc.pop("_id", None)

    admins = users_collection.find({"role": "admin"})
    async for adm in admins:
        await create_notification(
            user_id=adm["id"],
            title="New Website Lead",
            body=f"{data.full_name} ({data.phone}) — {data.service or 'General enquiry'}",
            n_type="lead",
            link="/portal/admin/leads",
        )
    return {"status": "received", "id": doc["id"]}


@router.post("/public/career")
async def capture_career_lead(data: PublicCareerLeadCreate):
    """Public, unauthenticated endpoint. The Career page should POST here
    (in addition to the resume upload) so applications show up under
    Admin Portal -> Leads -> Career Leads."""
    doc = {
        "id": str(uuid.uuid4()),
        "full_name": data.full_name,
        "phone": data.phone,
        "email": data.email,
        "position": data.position,
        "experience": data.experience,
        "message": data.message,
        "resume_url": data.resume_url,
        "father_name": data.father_name,
        "mother_name": data.mother_name,
        "dob": data.dob,
        "gender": data.gender,
        "marital_status": data.marital_status,
        "alt_mobile": data.alt_mobile,
        "city": data.city,
        "state": data.state,
        "address": data.address,
        "current_company": data.current_company,
        "current_designation": data.current_designation,
        "current_salary": data.current_salary,
        "expected_salary": data.expected_salary,
        "qualification": data.qualification,
        "college": data.college,
        "passing_year": data.passing_year,
        "skills": data.skills,
        "other_skills": data.other_skills,
        "ref_name": data.ref_name,
        "why_join": data.why_join,
        "additional_info": data.additional_info,
        "photo_url": data.photo_url,
        "hear_about_us": data.hear_about_us,
        "hear_about_us_other": data.hear_about_us_other,
        "status": "new",  # new, shortlisted, interview, hired, rejected
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await career_leads_collection.insert_one(doc)
    doc.pop("_id", None)

    admins = users_collection.find({"role": "admin"})
    async for adm in admins:
        await create_notification(
            user_id=adm["id"],
            title="New Career Application",
            body=f"{data.full_name} applied for {data.position or 'a role'}",
            n_type="lead",
            link="/portal/admin/leads",
        )

    if data.email:
        try:
            send_career_application_received_email(data.email, data.full_name, data.position)
        except Exception:
            logger.exception("Failed to send career application confirmation email to %s", data.email)

    return {"status": "received", "id": doc["id"]}


@router.put("/career/{career_lead_id}/status")
async def update_career_lead_status(career_lead_id: str, status: str = Query(...), admin: dict = Depends(require_admin)):
    valid = {"new", "shortlisted", "interview", "hired", "rejected"}
    if status not in valid:
        raise HTTPException(status_code=400, detail=f"Invalid status. Use: {', '.join(valid)}")
    result = await career_leads_collection.find_one_and_update(
        {"id": career_lead_id}, {"$set": {"status": status}}, return_document=True,
    )
    if not result:
        raise HTTPException(status_code=404, detail="Not found")
    result.pop("_id", None)
    return result


@router.delete("/career/{career_lead_id}")
async def delete_career_lead(career_lead_id: str, admin: dict = Depends(require_admin)):
    result = await career_leads_collection.delete_one({"id": career_lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"status": "deleted"}


# ── Proposal download tracking (public, no auth) ──

class ProposalTrackData(BaseModel):
    name: str = ""
    phone: str = ""
    source: str = "website_proposal"
    created_at: Optional[str] = None


@router.post("/public/proposal-track")
async def track_proposal_download(data: ProposalTrackData):
    """Track when someone downloads a proposal from the website calculator.
    Saves to web_leads so admin can see in Website Leads section."""
    doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "phone": data.phone,
        "source": data.source,
        "type": "proposal_download",
        "status": "new",
        "created_at": data.created_at or datetime.now(timezone.utc).isoformat(),
    }
    await web_leads_collection.insert_one(doc)
    doc.pop("_id", None)

    # Notify admin
    admins = users_collection.find({"role": "admin"})
    async for adm in admins:
        await create_notification(
            user_id=adm["id"],
            title="New Proposal Download",
            body=f"{data.name or 'Visitor'} ({data.phone}) downloaded a proposal",
            n_type="lead",
            link="/portal/admin/leads",
        )
    return {"status": "tracked", "id": doc["id"]}
